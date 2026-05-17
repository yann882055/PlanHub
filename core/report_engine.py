"""
core/report_engine.py — PlanHub v1.0
Génération de rapports (Excel, HTML, PDF) depuis les données XER.
"""

import os
import io
import datetime
from typing import List, Dict, Any, Optional


class ReportEngine:
    """Moteur de génération de rapports d'avancement."""

    def __init__(self):
        self.project_info: Dict = {}
        self.header_info: Dict = {}

    def set_header(
        self,
        project_name: str = "",
        report_date: str = "",
        title: str = "",
        planner: str = "",
        revision: str = "Rev 0",
        logo_path: str = "",
    ):
        """Définit l'en-tête du rapport."""
        self.header_info = {
            "project_name": project_name,
            "report_date": report_date or datetime.date.today().strftime("%d/%m/%Y"),
            "title": title or "Rapport d'avancement",
            "planner": planner,
            "revision": revision,
            "logo_path": logo_path,
        }

    def export_excel(
        self,
        rows: List[Dict],
        columns: List[str],
        filepath: str,
    ) -> bool:
        """
        Exporte les données en Excel (.xlsx).
        Nécessite openpyxl.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Planning"

            # Couleurs
            BLUE = "1565C0"
            LIGHT_BLUE = "E3F2FD"
            RED = "D32F2F"
            ORANGE = "FF9800"
            GRAY = "F5F5F5"

            header_fill = PatternFill("solid", fgColor=BLUE)
            header_font = Font(bold=True, color="FFFFFF", name="Segoe UI", size=9)
            data_font = Font(name="Segoe UI", size=9)
            critical_fill = PatternFill("solid", fgColor="FFEBEE")

            # En-tête rapport
            row_num = 1
            if self.header_info:
                ws.merge_cells(f"A{row_num}:{get_column_letter(len(columns))}{row_num}")
                ws[f"A{row_num}"] = self.header_info.get("project_name", "")
                ws[f"A{row_num}"].font = Font(bold=True, size=14, name="Segoe UI", color=BLUE)
                row_num += 1

                ws.merge_cells(f"A{row_num}:{get_column_letter(len(columns))}{row_num}")
                ws[f"A{row_num}"] = (
                    f"{self.header_info.get('title','')} | "
                    f"{self.header_info.get('report_date','')} | "
                    f"{self.header_info.get('revision','')}"
                )
                ws[f"A{row_num}"].font = Font(size=10, name="Segoe UI", color="616161")
                row_num += 2

            # En-tête colonnes
            COLUMN_LABELS = {
                "activity_id": "Activity ID",
                "wbs_code": "WBS",
                "wbs_name": "Nom WBS",
                "activity_name": "Désignation",
                "orig_dur": "Durée orig.",
                "remain_dur": "Durée reste",
                "act_dur": "Durée réelle",
                "start": "Début",
                "finish": "Fin",
                "early_start": "Début tôt",
                "early_finish": "Fin tôt",
                "late_start": "Début tard",
                "late_finish": "Fin tard",
                "total_float": "Marge totale",
                "free_float": "Marge libre",
                "budgeted_cost": "Coût budgété",
                "actual_cost": "Coût réel",
                "remain_cost": "Coût restant",
                "pct_complete": "% Avancement",
                "cstr_type": "Contrainte",
                "cstr_date": "Date contrainte",
                "status": "Statut",
                "predecessors": "Prédécesseurs",
                "project_id": "Projet ID",
            }

            for col_idx, col_key in enumerate(columns, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = COLUMN_LABELS.get(col_key, col_key)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.row_dimensions[row_num].height = 30

            row_num += 1

            # Données
            for row_data in rows:
                is_critical = row_data.get("_is_critical", False)
                for col_idx, col_key in enumerate(columns, start=1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    value = row_data.get(col_key, "")
                    try:
                        if col_key in ("orig_dur", "remain_dur", "act_dur",
                                       "total_float", "free_float"):
                            cell.value = int(value) if value != "" else ""
                        elif col_key in ("budgeted_cost", "actual_cost", "remain_cost"):
                            cell.value = float(value) if value != "" else 0
                            cell.number_format = "#,##0"
                        elif col_key == "pct_complete":
                            cell.value = float(value) / 100 if value != "" else 0
                            cell.number_format = "0%"
                        else:
                            cell.value = str(value) if value else ""
                    except Exception:
                        cell.value = str(value) if value else ""

                    cell.font = data_font
                    if is_critical:
                        cell.fill = PatternFill("solid", fgColor="FFCDD2")
                    elif row_num % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor=GRAY)

                row_num += 1

            # Largeurs colonnes
            col_widths = {
                "activity_id": 12, "wbs_code": 10, "wbs_name": 25,
                "activity_name": 40, "orig_dur": 10, "remain_dur": 10,
                "start": 16, "finish": 16, "early_start": 16, "early_finish": 16,
                "late_start": 16, "late_finish": 16,
                "total_float": 10, "free_float": 10,
                "budgeted_cost": 15, "pct_complete": 12,
                "predecessors": 20,
            }
            for col_idx, col_key in enumerate(columns, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_key, 15)

            # Figer première ligne après en-tête
            ws.freeze_panes = f"A{row_num - len(rows)}"

            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Erreur export Excel : {e}")
            return False

    def export_html(
        self,
        rows: List[Dict],
        columns: List[str],
        filepath: str,
    ) -> bool:
        """Exporte en HTML standalone imprimable."""
        try:
            COLUMN_LABELS = {
                "activity_id": "Activity ID", "wbs_code": "WBS",
                "activity_name": "Désignation", "orig_dur": "Durée",
                "start": "Début", "finish": "Fin",
                "total_float": "Marge", "budgeted_cost": "Coût budgété",
                "pct_complete": "%", "status": "Statut",
                "predecessors": "Prédécesseurs",
            }

            h = self.header_info
            html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<title>{h.get('title','Planning')} — {h.get('project_name','')}</title>
<style>
  body{{font-family:'Segoe UI',Arial,sans-serif;font-size:11px;margin:20px;color:#212121}}
  h1{{color:#1565C0;font-size:16px;margin:0}}
  .subtitle{{color:#616161;font-size:11px;margin:4px 0 16px}}
  table{{width:100%;border-collapse:collapse;margin-top:12px}}
  th{{background:#1565C0;color:#fff;padding:6px 8px;text-align:left;font-size:10px;white-space:nowrap}}
  td{{padding:5px 8px;border-bottom:1px solid #E0E0E0;white-space:nowrap}}
  tr:nth-child(even) td{{background:#F5F5F5}}
  tr.critical td{{background:#FFEBEE}}
  tr.milestone td{{background:#E8F5E9;font-weight:bold}}
  @media print{{body{{margin:10mm}} th{{font-size:8px}} td{{font-size:8px}}}}
</style>
</head>
<body>
<h1>{h.get('project_name','Projet')}</h1>
<div class="subtitle">
  {h.get('title','Rapport')} &mdash; {h.get('report_date','')} &mdash; {h.get('revision','')}
  {' &mdash; Planificateur : ' + h.get('planner','') if h.get('planner') else ''}
</div>
<table>
<thead><tr>
"""
            for col in columns:
                html += f"<th>{COLUMN_LABELS.get(col, col)}</th>"
            html += "</tr></thead><tbody>\n"

            for row in rows:
                cls = "critical" if row.get("_is_critical") else (
                    "milestone" if row.get("_is_milestone") else ""
                )
                html += f'<tr class="{cls}">'
                for col in columns:
                    val = row.get(col, "")
                    html += f"<td>{val}</td>"
                html += "</tr>\n"

            html += "</tbody></table>\n</body></html>"

            with open(filepath, "w", encoding="utf-8") as f:
                f.write(html)
            return True
        except Exception as e:
            print(f"Erreur export HTML : {e}")
            return False

    def export_pdf(
        self,
        rows: List[Dict],
        columns: List[str],
        filepath: str,
        orientation: str = "landscape",
    ) -> bool:
        """Exporte en PDF via reportlab."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, A3, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm

            pagesize = landscape(A4) if orientation == "landscape" else A4
            doc = SimpleDocTemplate(
                filepath,
                pagesize=pagesize,
                topMargin=15*mm, bottomMargin=15*mm,
                leftMargin=15*mm, rightMargin=15*mm
            )

            styles = getSampleStyleSheet()
            BLUE = colors.HexColor("#1565C0")

            story = []

            # En-tête
            h = self.header_info
            title_style = ParagraphStyle("title", parent=styles["Normal"],
                                          fontSize=14, textColor=BLUE, fontName="Helvetica-Bold")
            story.append(Paragraph(h.get("project_name", "Projet"), title_style))
            story.append(Paragraph(
                f"{h.get('title','')} | {h.get('report_date','')} | {h.get('revision','')}",
                styles["Normal"]
            ))
            story.append(Spacer(1, 10*mm))

            COLUMN_LABELS = {
                "activity_id": "Act. ID", "wbs_code": "WBS",
                "activity_name": "Désignation", "orig_dur": "Durée",
                "start": "Début", "finish": "Fin",
                "total_float": "Marge", "budgeted_cost": "Coût",
                "pct_complete": "%", "predecessors": "Préd.",
            }

            # Table
            header_row = [COLUMN_LABELS.get(c, c) for c in columns]
            data = [header_row]

            for row in rows:
                data_row = []
                for col in columns:
                    val = row.get(col, "")
                    if col == "budgeted_cost":
                        try:
                            val = f"{float(val):,.0f}"
                        except Exception:
                            pass
                    data_row.append(str(val) if val else "")
                data.append(data_row)

            col_widths = [20*mm if c == "activity_id" else
                          35*mm if c == "activity_name" else
                          15*mm for c in columns]

            table = Table(data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), BLUE),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E0E0E0")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("PADDING", (0, 0), (-1, -1), 3),
            ]))

            story.append(table)
            doc.build(story)
            return True
        except Exception as e:
            print(f"Erreur export PDF : {e}")
            return False
