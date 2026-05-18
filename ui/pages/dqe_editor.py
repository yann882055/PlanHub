"""
dqe_editor.py - Éditeur DQE (Devis Quantitatif Estimatif) complet
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
from datetime import datetime

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# Colonnes du tableau DQE (calendrier et contrainte gérés en interne, pas affichés)
DQE_COLUMNS = [
    ("activity_id", "N°",          60,  "center"),
    ("wbs",         "WBS",         90,  "center"),
    ("lot",         "LOT",         90,  "center"),
    ("designation", "DÉSIGNATION", 260, "w"),
    ("unite",       "UNITÉ",       80,  "center"),
    ("quantite",    "QUANTITÉ",    90,  "e"),
    ("pu_ht",       "PU HT",       110, "e"),
    ("montant_ht",  "MONTANT HT",  120, "e"),
    ("duree",       "DURÉE (j)",   90,  "center"),
    ("task_type",   "TYPE",        90,  "center"),
]

UNITES = ["m³", "m²", "ml", "forfait", "kg", "T", "U", "h", "j"]
TASK_TYPES = ["TT_Task", "TT_Mile", "TT_FinMile", "TT_LOE", "TT_WBS"]
DEVISES = ["FCFA", "XOF", "USD", "EUR"]


class DQEEditorPage(ctk.CTkFrame):
    """Éditeur DQE avec Treeview éditable, import/export Excel."""

    def __init__(self, parent, project_state: dict, navigate, update_status):
        super().__init__(parent, fg_color="white")
        self.project_state = project_state
        self.navigate = navigate
        self.update_status = update_status

        self._edit_widget = None
        self._edit_col = None
        self._edit_item = None
        self._next_id = 1

        self._build_ui()

    # ------------------------------------------------------------------
    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)

        # En-tête
        self._build_header()
        # Barre d'outils
        self._build_toolbar()
        # Tableau
        self._build_table()
        # Pied de page (total)
        self._build_footer()

    def _build_header(self):
        hdr = ctk.CTkFrame(self, fg_color="#1565C0", corner_radius=0)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.columnconfigure(1, weight=1)

        ctk.CTkLabel(
            hdr, text="📋  Éditeur DQE",
            font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"),
            text_color="white"
        ).grid(row=0, column=0, padx=20, pady=14, sticky="w")

        ctk.CTkLabel(
            hdr, text="PlanHub  ›  DQE",
            font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color="#BBDEFB"
        ).grid(row=0, column=1, padx=10, pady=14, sticky="e")

    def _build_toolbar(self):
        bar = ctk.CTkFrame(self, fg_color="#F5F5F5", corner_radius=0, border_width=1, border_color="#E0E0E0")
        bar.grid(row=1, column=0, sticky="ew")

        btn_cfg = dict(height=32, corner_radius=6, font=ctk.CTkFont(family="Segoe UI", size=11))

        ctk.CTkButton(bar, text="➕ Ajouter", fg_color="#1565C0", hover_color="#0D47A1",
                      text_color="white", command=self._add_row, **btn_cfg
                      ).pack(side="left", padx=(10, 4), pady=8)

        ctk.CTkButton(bar, text="🗑 Supprimer", fg_color="#C62828", hover_color="#B71C1C",
                      text_color="white", command=self._delete_row, **btn_cfg
                      ).pack(side="left", padx=4, pady=8)

        ctk.CTkButton(bar, text="⎘ Dupliquer", fg_color="#37474F", hover_color="#263238",
                      text_color="white", command=self._duplicate_row, **btn_cfg
                      ).pack(side="left", padx=4, pady=8)

        ctk.CTkButton(bar, text="▲ Monter", fg_color="#37474F", hover_color="#263238",
                      text_color="white", command=self._move_up, **btn_cfg
                      ).pack(side="left", padx=4, pady=8)

        ctk.CTkButton(bar, text="▼ Descendre", fg_color="#37474F", hover_color="#263238",
                      text_color="white", command=self._move_down, **btn_cfg
                      ).pack(side="left", padx=4, pady=8)

        # Séparateur
        ttk.Separator(bar, orient="vertical").pack(side="left", fill="y", padx=8, pady=6)

        ctk.CTkButton(bar, text="📥 Import Excel", fg_color="#2E7D32", hover_color="#1B5E20",
                      text_color="white", command=self._import_excel, **btn_cfg
                      ).pack(side="left", padx=4, pady=8)

        ctk.CTkButton(bar, text="📤 Export Excel", fg_color="#2E7D32", hover_color="#1B5E20",
                      text_color="white", command=self._export_excel, **btn_cfg
                      ).pack(side="left", padx=4, pady=8)

        # Devise
        ctk.CTkLabel(bar, text="Devise :", font=ctk.CTkFont(family="Segoe UI", size=11),
                     text_color="#333333"
                     ).pack(side="right", padx=(4, 0), pady=8)

        self.currency_var = tk.StringVar(value=self.project_state.get("currency", "FCFA"))
        currency_cb = ctk.CTkComboBox(
            bar, values=DEVISES, variable=self.currency_var, width=90, height=32,
            command=self._on_currency_change
        )
        currency_cb.pack(side="right", padx=(4, 12), pady=8)

    def _build_table(self):
        table_frame = ctk.CTkFrame(self, fg_color="white")
        table_frame.grid(row=2, column=0, sticky="nsew", padx=0, pady=0)
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("DQE.Treeview", background="white", fieldbackground="white",
                        rowheight=26, font=("Segoe UI", 10))
        style.configure("DQE.Treeview.Heading", background="#1565C0", foreground="white",
                        font=("Segoe UI", 10, "bold"))
        style.map("DQE.Treeview", background=[("selected", "#BBDEFB")], foreground=[("selected", "#0D47A1")])

        col_ids = [c[0] for c in DQE_COLUMNS]
        self.tv = ttk.Treeview(table_frame, columns=col_ids, show="headings", style="DQE.Treeview")

        for col_id, col_lbl, col_w, col_anchor in DQE_COLUMNS:
            self.tv.heading(col_id, text=col_lbl, command=lambda c=col_id: self._sort_column(c))
            self.tv.column(col_id, width=col_w, anchor=col_anchor, minwidth=50)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tv.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tv.xview)
        self.tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        self.tv.bind("<Double-1>", self._on_double_click)
        self.tv.bind("<Button-1>", self._cancel_edit)
        self.tv.tag_configure("odd", background="#FAFAFA")
        self.tv.tag_configure("even", background="white")

    def _build_footer(self):
        footer = ctk.CTkFrame(self, fg_color="#ECEFF1", corner_radius=0, border_width=1, border_color="#CFD8DC")
        footer.grid(row=3, column=0, sticky="ew")

        ctk.CTkLabel(footer, text="TOTAL GÉNÉRAL :", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
                     text_color="#1565C0"
                     ).pack(side="left", padx=16, pady=8)

        self.total_var = tk.StringVar(value="0 FCFA")
        ctk.CTkLabel(footer, textvariable=self.total_var,
                     font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
                     text_color="#1B5E20"
                     ).pack(side="left", padx=4, pady=8)

        self.count_var = tk.StringVar(value="0 ligne(s)")
        ctk.CTkLabel(footer, textvariable=self.count_var,
                     font=ctk.CTkFont(family="Segoe UI", size=11),
                     text_color="#555555"
                     ).pack(side="right", padx=16, pady=8)

    # ------------------------------------------------------------------
    # Actions
    # ------------------------------------------------------------------

    def _next_activity_id(self):
        existing = [t.get("activity_id", "") for t in self.project_state.get("dqe_tasks", [])]
        pid = self.project_state.get("proj_id", "PRJ") or "PRJ"
        idx = 1
        while f"A{idx:04d}" in existing:
            idx += 1
        return f"A{idx:04d}"

    def _add_row(self):
        new_task = {
            "activity_id": self._next_activity_id(),
            "wbs":         "1",
            "lot":         "",
            "designation": "Nouvelle tâche",
            "unite":       "forfait",
            "quantite":    "1",
            "pu_ht":       "0",
            "montant_ht":  "0",
            "duree":       "1",
            "calendrier":  "Cal_5j",
            "task_type":   "TT_Task",
            "constraint":  "CS_ALAP",
        }
        self.project_state.setdefault("dqe_tasks", []).append(new_task)
        self._reload_table()
        self.update_status("Ligne ajoutée")

    def _delete_row(self):
        selected = self.tv.selection()
        if not selected:
            messagebox.showwarning("Sélection", "Veuillez sélectionner une ligne à supprimer.")
            return
        if not messagebox.askyesno("Confirmation", f"Supprimer {len(selected)} ligne(s) sélectionnée(s) ?"):
            return
        ids_to_del = {self.tv.item(i, "values")[0] for i in selected}
        tasks = [t for t in self.project_state.get("dqe_tasks", [])
                 if t.get("activity_id") not in ids_to_del]
        self.project_state["dqe_tasks"] = tasks
        self._reload_table()
        self.update_status(f"{len(ids_to_del)} ligne(s) supprimée(s)")

    def _duplicate_row(self):
        selected = self.tv.selection()
        if not selected:
            return
        item = selected[0]
        idx = self.tv.index(item)
        tasks = self.project_state.get("dqe_tasks", [])
        if idx < len(tasks):
            import copy
            dup = copy.deepcopy(tasks[idx])
            dup["activity_id"] = self._next_activity_id()
            dup["designation"] = dup["designation"] + " (copie)"
            tasks.insert(idx + 1, dup)
            self.project_state["dqe_tasks"] = tasks
            self._reload_table()

    def _move_up(self):
        selected = self.tv.selection()
        if not selected:
            return
        idx = self.tv.index(selected[0])
        tasks = self.project_state.get("dqe_tasks", [])
        if idx > 0:
            tasks[idx - 1], tasks[idx] = tasks[idx], tasks[idx - 1]
            self._reload_table()
            # Re-sélectionner
            items = self.tv.get_children()
            if idx - 1 < len(items):
                self.tv.selection_set(items[idx - 1])

    def _move_down(self):
        selected = self.tv.selection()
        if not selected:
            return
        idx = self.tv.index(selected[0])
        tasks = self.project_state.get("dqe_tasks", [])
        if idx < len(tasks) - 1:
            tasks[idx], tasks[idx + 1] = tasks[idx + 1], tasks[idx]
            self._reload_table()
            items = self.tv.get_children()
            if idx + 1 < len(items):
                self.tv.selection_set(items[idx + 1])

    def _on_currency_change(self, value):
        self.project_state["currency"] = value
        self._update_total()

    def _sort_column(self, col):
        """Tri par colonne."""
        tasks = self.project_state.get("dqe_tasks", [])
        try:
            tasks.sort(key=lambda t: float(t.get(col, 0) or 0))
        except (ValueError, TypeError):
            tasks.sort(key=lambda t: str(t.get(col, "")))
        self._reload_table()

    # ------------------------------------------------------------------
    # Édition inline
    # ------------------------------------------------------------------

    def _on_double_click(self, event):
        self._cancel_edit(event)
        region = self.tv.identify("region", event.x, event.y)
        if region != "cell":
            return

        col_id = self.tv.identify_column(event.x)
        item = self.tv.identify_row(event.y)
        if not item:
            return

        col_index = int(col_id.replace("#", "")) - 1
        col_name = DQE_COLUMNS[col_index][0]

        # Colonnes avec liste de choix
        choice_cols = {
            "unite": UNITES,
            "task_type": TASK_TYPES,
        }

        bbox = self.tv.bbox(item, col_id)
        if not bbox:
            return

        current_val = self.tv.item(item, "values")[col_index]

        if col_name in choice_cols:
            var = tk.StringVar(value=current_val)
            cb = ttk.Combobox(self.tv, textvariable=var, values=choice_cols[col_name], state="readonly")
            cb.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
            cb.focus()
            cb.bind("<<ComboboxSelected>>", lambda e, i=item, c=col_name, ci=col_index, w=cb, v=var:
                    self._save_combobox(i, c, ci, w, v))
            cb.bind("<Escape>", lambda e, w=cb: w.destroy())
            self._edit_widget = cb
        else:
            var = tk.StringVar(value=current_val)
            entry = tk.Entry(self.tv, textvariable=var, font=("Segoe UI", 10))
            entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
            entry.focus()
            entry.select_range(0, "end")
            entry.bind("<Return>", lambda e, i=item, c=col_name, ci=col_index, w=entry, v=var:
                       self._save_entry(i, c, ci, w, v))
            entry.bind("<Escape>", lambda e, w=entry: w.destroy())
            entry.bind("<Tab>", lambda e, i=item, c=col_name, ci=col_index, w=entry, v=var:
                       self._save_entry(i, c, ci, w, v))
            self._edit_widget = entry

        self._edit_item = item
        self._edit_col = col_name

    def _save_entry(self, item, col_name, col_index, widget, var):
        value = var.get()
        widget.destroy()
        self._edit_widget = None
        self._update_task_field(item, col_name, value)

    def _save_combobox(self, item, col_name, col_index, widget, var):
        value = var.get()
        widget.destroy()
        self._edit_widget = None
        self._update_task_field(item, col_name, value)

    def _cancel_edit(self, event=None):
        if self._edit_widget:
            try:
                self._edit_widget.destroy()
            except Exception:
                pass
            self._edit_widget = None

    def _update_task_field(self, item, col_name, value):
        idx = self.tv.index(item)
        tasks = self.project_state.get("dqe_tasks", [])
        if idx < len(tasks):
            tasks[idx][col_name] = value
            # Recalcul montant
            if col_name in ("quantite", "pu_ht"):
                try:
                    q = float(tasks[idx].get("quantite", 0) or 0)
                    p = float(tasks[idx].get("pu_ht", 0) or 0)
                    tasks[idx]["montant_ht"] = str(round(q * p, 2))
                except ValueError:
                    pass
            self._reload_table()

    # ------------------------------------------------------------------
    # Import / Export Excel
    # ------------------------------------------------------------------

    # ------------------------------------------------------------------
    # Détection intelligente du format Excel
    # ------------------------------------------------------------------

    def _detect_header_row(self, ws):
        """Trouve la ligne d'en-tête et le mapping colonnes.
        Retourne (header_row_idx, col_poste, col_desig, col_montant, col_qte, col_pu, col_unite, col_duree, devise)
        """
        POSTE_KEYS  = {"POSTE", "N°", "N°", "NO", "CODE", "WBS", "ID", "REF"}
        DESIG_KEYS  = {"DÉSIGNATION", "DESIGNATION", "LIBELLÉ", "LIBELLE",
                       "DESCRIPTION", "INTITULÉ", "INTITULE", "NATURE"}
        MONTANT_KEYS= {"MONTANT", "TOTAL", "PRIX TOTAL", "MONTANT HT", "MONTANT TOTAL",
                       "PRIX", "AMOUNT"}
        QTE_KEYS    = {"QUANTITÉ", "QUANTITE", "QTE", "QTÉ"}
        PU_KEYS     = {"PU HT", "PUHT", "PU", "PRIX UNIT", "PRIX UNITAIRE", "UNIT PRICE"}
        UNITE_KEYS  = {"UNITÉ", "UNITE", "UNI", "UNIT"}
        DUREE_KEYS  = {"DURÉE", "DUREE", "DURATION", "DÉLAI", "DELAI"}

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
            headers = [str(v or "").strip().upper().replace("\n", " ") for v in row]
            # Cherche au moins Désignation + (Poste ou Montant)
            has_desig  = any(any(k in h for k in DESIG_KEYS)  for h in headers)
            has_poste  = any(any(k in h for k in POSTE_KEYS)  for h in headers)
            has_montant= any(any(k in h for k in MONTANT_KEYS) for h in headers)
            if has_desig and (has_poste or has_montant):
                def find_col(keys):
                    for i, h in enumerate(headers):
                        if any(k in h for k in keys):
                            return i
                    return None
                # Deviner la devise dans les entêtes
                devise = "FCFA"
                for h in headers:
                    if "USD" in h:  devise = "USD"; break
                    if "EUR" in h:  devise = "EUR"; break
                    if "XOF" in h:  devise = "XOF"; break
                return (row_idx,
                        find_col(POSTE_KEYS), find_col(DESIG_KEYS),
                        find_col(MONTANT_KEYS), find_col(QTE_KEYS),
                        find_col(PU_KEYS), find_col(UNITE_KEYS),
                        find_col(DUREE_KEYS), devise)
        return None

    def _is_wbs_code(self, code: str) -> bool:
        """Retourne True si c'est un code WBS numérique (1, 1.1, 2.3.1)."""
        import re
        return bool(re.match(r'^\d+(\.\d+)*$', str(code).strip()))

    def _import_excel(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("Module manquant",
                "openpyxl n'est pas installé.\nInstallez-le avec : pip install openpyxl")
            return

        path = filedialog.askopenfilename(
            title="Importer un fichier DQE Excel",
            filetypes=[("Fichiers Excel", "*.xlsx *.xls"), ("Tous", "*.*")]
        )
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path, data_only=True)

            # Choisir la feuille
            sheet_names = wb.sheetnames
            ws = wb.active
            if len(sheet_names) > 1:
                # Prendre la première feuille qui ressemble à un DQE
                for sn in sheet_names:
                    info = self._detect_header_row(wb[sn])
                    if info:
                        ws = wb[sn]
                        break

            info = self._detect_header_row(ws)

            # ── Cas 1 : entête standard détecté ──────────────────────────
            if info:
                (hrow, ci_poste, ci_desig, ci_mont,
                 ci_qte, ci_pu, ci_unite, ci_duree, devise) = info

                # Mettre à jour la devise si différente
                if devise != self.currency_var.get():
                    self.currency_var.set(devise)
                    self.project_state["currency"] = devise

                new_tasks = []
                current_lot = ""
                act_counter = [1]

                def next_id():
                    while f"A{act_counter[0]:04d}" in \
                          [t.get("activity_id","") for t in self.project_state.get("dqe_tasks",[])]:
                        act_counter[0] += 1
                    aid = f"A{act_counter[0]:04d}"
                    act_counter[0] += 1
                    return aid

                for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
                    if all(v is None for v in row):
                        continue

                    def cell(idx):
                        if idx is None or idx >= len(row):
                            return ""
                        return row[idx]

                    poste   = str(cell(ci_poste)  or "").strip()
                    desig   = str(cell(ci_desig)  or "").strip()
                    montant = cell(ci_mont)
                    qte     = cell(ci_qte)
                    pu      = cell(ci_pu)
                    unite   = str(cell(ci_unite)  or "").strip() or "forfait"
                    duree   = cell(ci_duree)

                    if not desig and not poste:
                        continue
                    # Ignorer les lignes sans code Poste et sans montant (métadonnées/jalons vides)
                    if not poste and not montant:
                        continue

                    # Déterminer type
                    is_wbs = self._is_wbs_code(poste) if poste else False
                    task_type = "TT_WBS" if is_wbs else "TT_Task"

                    # Gérer le lot courant
                    if is_wbs:
                        current_lot = poste

                    # Montant
                    try:
                        mont_val = round(float(str(montant).replace(" ", "").replace(",", ".") or 0), 2)
                    except Exception:
                        mont_val = 0.0

                    # PU / Quantité
                    try:
                        qte_val = float(str(qte).replace(" ", "").replace(",", ".") or 1)
                    except Exception:
                        qte_val = 1.0
                    try:
                        pu_val = float(str(pu).replace(" ", "").replace(",", ".") or 0)
                    except Exception:
                        pu_val = 0.0

                    # Si montant absent mais PU et QTE présents
                    if mont_val == 0 and pu_val and qte_val:
                        mont_val = round(pu_val * qte_val, 2)

                    # Durée
                    try:
                        duree_val = int(float(str(duree or "").replace(",", ".") or 1))
                    except Exception:
                        duree_val = 1

                    # Activity ID
                    if poste and not is_wbs:
                        activity_id = poste          # Ex: A100, A101
                    elif poste and is_wbs:
                        activity_id = f"WBS_{poste.replace('.','_')}"
                    else:
                        activity_id = next_id()

                    task = {
                        "activity_id": activity_id,
                        "wbs":         poste or current_lot,
                        "lot":         current_lot if not is_wbs else "",
                        "designation": desig,
                        "unite":       unite if unite in UNITES else "forfait",
                        "quantite":    str(qte_val) if qte_val != 1.0 else "1",
                        "pu_ht":       str(pu_val) if pu_val else "",
                        "montant_ht":  str(mont_val) if mont_val else "",
                        "duree":       str(duree_val) if not is_wbs else "",
                        "calendrier":  "Cal_5j",
                        "task_type":   task_type,
                        "constraint":  "CS_ALAP",
                    }
                    new_tasks.append(task)

            # ── Cas 2 : format libre — 2-3 colonnes sans entête reconnu ──
            else:
                new_tasks = []
                current_lot = ""
                act_counter = [1]

                def next_id2():
                    aid = f"A{act_counter[0]:04d}"
                    act_counter[0] += 1
                    return aid

                # Chercher la première ligne avec données utiles (≥ 2 cellules non vides)
                data_start = 1
                for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
                    non_empty = [v for v in row if v is not None]
                    if len(non_empty) >= 2 and any(
                            isinstance(non_empty[0], str) and non_empty[0].strip()):
                        data_start = ri
                        break

                devise = "FCFA"
                for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                    for v in row:
                        sv = str(v or "").upper()
                        if "USD" in sv: devise = "USD"; break
                        if "EUR" in sv: devise = "EUR"; break

                if devise != self.currency_var.get():
                    self.currency_var.set(devise)
                    self.project_state["currency"] = devise

                for row in ws.iter_rows(min_row=data_start, values_only=True):
                    vals = list(row)
                    if all(v is None for v in vals):
                        continue

                    # Col 0 = Poste, Col 1 = Désignation, Col 2 = Montant
                    poste   = str(vals[0] or "").strip() if len(vals) > 0 else ""
                    desig   = str(vals[1] or "").strip() if len(vals) > 1 else ""
                    montant = vals[2] if len(vals) > 2 else None

                    if not desig:
                        continue

                    is_wbs = self._is_wbs_code(poste) if poste else False
                    if is_wbs:
                        current_lot = poste

                    try:
                        mont_val = round(float(str(montant or "").replace(" ","").replace(",",".")), 2)
                    except Exception:
                        mont_val = 0.0

                    if poste and not is_wbs:
                        activity_id = poste
                    elif poste and is_wbs:
                        activity_id = f"WBS_{poste.replace('.','_')}"
                    else:
                        activity_id = next_id2()

                    task = {
                        "activity_id": activity_id,
                        "wbs":         poste or current_lot,
                        "lot":         current_lot if not is_wbs else "",
                        "designation": desig,
                        "unite":       "forfait",
                        "quantite":    "1",
                        "pu_ht":       "",
                        "montant_ht":  str(mont_val) if mont_val else "",
                        "duree":       "1" if not is_wbs else "",
                        "calendrier":  "Cal_5j",
                        "task_type":   "TT_WBS" if is_wbs else "TT_Task",
                        "constraint":  "CS_ALAP",
                    }
                    new_tasks.append(task)

            # ── Intégration dans le projet ────────────────────────────────
            if new_tasks:
                n = len(new_tasks)
                rep = messagebox.askyesno(
                    "Import DQE",
                    f"{n} lignes importées ({devise}).\n\n"
                    "Cliquez OUI pour remplacer le DQE existant.\n"
                    "Cliquez NON pour ajouter à la suite.",
                    icon="question"
                )
                if rep:
                    self.project_state["dqe_tasks"] = new_tasks
                else:
                    self.project_state.setdefault("dqe_tasks", []).extend(new_tasks)
                self._reload_table()
                self.update_status(f"Import Excel : {n} ligne(s) importée(s) — {devise}")
            else:
                messagebox.showinfo("Import", "Aucune donnée trouvée dans le fichier.")

        except Exception as ex:
            messagebox.showerror("Erreur import", f"Impossible d'importer le fichier :\n{ex}")

    def _export_excel(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("Module manquant", "openpyxl n'est pas installé.")
            return

        path = filedialog.asksaveasfilename(
            title="Exporter le DQE",
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")],
            initialfile=f"DQE_{self.project_state.get('name', 'projet')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "DQE"

            # En-têtes
            headers = [col_lbl for _, col_lbl, _, _ in DQE_COLUMNS]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="1565C0")

            # Données
            tasks = self.project_state.get("dqe_tasks", [])
            for r, task in enumerate(tasks, 2):
                for c, (col_id, _, _, _) in enumerate(DQE_COLUMNS, 1):
                    ws.cell(row=r, column=c, value=task.get(col_id, ""))

            # Ligne total
            r_total = len(tasks) + 2
            ws.cell(row=r_total, column=1, value="TOTAL").font = openpyxl.styles.Font(bold=True)
            total = sum(float(t.get("montant_ht", 0) or 0) for t in tasks)
            col_m = next((i + 1 for i, (cid, _, _, _) in enumerate(DQE_COLUMNS) if cid == "montant_ht"), 8)
            cell_t = ws.cell(row=r_total, column=col_m, value=round(total, 2))
            cell_t.font = openpyxl.styles.Font(bold=True)

            wb.save(path)
            messagebox.showinfo("Export", f"DQE exporté avec succès :\n{path}")
            self.update_status(f"Export Excel : {path}")

        except Exception as ex:
            messagebox.showerror("Erreur export", str(ex))

    # ------------------------------------------------------------------
    # Rechargement
    # ------------------------------------------------------------------

    def _reload_table(self):
        for item in self.tv.get_children():
            self.tv.delete(item)

        tasks = self.project_state.get("dqe_tasks", [])
        for i, task in enumerate(tasks):
            values = tuple(task.get(col_id, "") for col_id, _, _, _ in DQE_COLUMNS)
            tag = "even" if i % 2 == 0 else "odd"
            self.tv.insert("", "end", values=values, tags=(tag,))

        self._update_total()
        self.count_var.set(f"{len(tasks)} ligne(s)")

    def _update_total(self):
        tasks = self.project_state.get("dqe_tasks", [])
        total = sum(float(t.get("montant_ht", 0) or 0) for t in tasks)
        currency = self.currency_var.get()
        self.total_var.set(f"{total:,.2f} {currency}")

    def refresh(self):
        self.currency_var.set(self.project_state.get("currency", "FCFA"))
        self._reload_table()
        self.update_status("Éditeur DQE rechargé")
